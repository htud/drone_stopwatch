import serial
import pandas as pd
from tabulate import tabulate
import os
from openpyxl import load_workbook

# 시리얼 포트 설정 (Arduino 포트와 보드 레이트에 맞게 설정)
ser = serial.Serial('COM8', 9600)  # 실제 포트 이름으로 변경하세요.

# 엑셀 파일 이름 설정
file_name = "timer_data.xlsx"

# 현재 작업 디렉토리 출력
current_directory = os.getcwd()
print(f"엑셀 파일 저장 위치: {current_directory}\\{file_name}")

# 데이터 저장을 위한 빈 데이터프레임 생성
data = {
    "팀": [],
    "A 타임": [],
    "B 타임": []
}

# 초기 팀 번호 설정
team_number = 1
time_a = None  # A 타이머 결과
time_b = None  # B 타이머 결과

print("엑셀 파일 초기화 완료: timer_data.xlsx")


# 현재 DataFrame을 테이블 형식으로 출력하는 함수
def print_table(df):
    print("\n현재 DataFrame 내용:")
    print(tabulate(df, headers='keys', tablefmt='pretty', showindex=False))


# 엑셀 파일에 데이터 추가 함수
def save_to_excel(df):
    if not os.path.exists(file_name):
        # 파일이 없는 경우 새로 생성
        with pd.ExcelWriter(file_name, mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name="Timer Data")
    else:
        # 파일이 있는 경우 마지막 줄에 추가
        with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # openpyxl로 기존 워크북 로드하여 마지막 행 찾아서 추가
            book = load_workbook(file_name)
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            startrow = writer.sheets["Timer Data"].max_row
            df.to_excel(writer, index=False, header=False, sheet_name="Timer Data", startrow=startrow)


try:
    while True:
        if ser.in_waiting > 0:
            # 시리얼 데이터 읽기
            line = ser.readline().decode('utf-8').strip()
            print(f"Received line: {line}")  # 수신된 라인 출력

            # A와 B 타이머 정지 결과 수신 및 저장
            if "A_FINAL:" in line or "B_FINAL:" in line:
                timer_label = "A" if "A_FINAL:" in line else "B"
                time_value = line.split(f"{timer_label}_FINAL:")[1].strip()

                # 각 타이머에 맞게 시간 값을 저장
                if timer_label == "A":
                    time_a = time_value
                else:
                    time_b = time_value

                print(f"{timer_label} 타이머 정지 - {timer_label} 타임: {time_value}")

            # 리셋 신호를 받았을 때 엑셀 파일에 기록
            elif "Sending reset signal" in line:
                if time_a is not None and time_b is not None:
                    # 데이터프레임에 새로운 데이터 추가
                    data["팀"].append(f"{team_number}팀")
                    data["A 타임"].append(time_a)
                    data["B 타임"].append(time_b)

                    # DataFrame 생성 후 엑셀에 저장
                    df = pd.DataFrame(data)
                    save_to_excel(df)

                    print(f"엑셀 파일에 저장 완료 - 팀: {team_number}팀, A 타임: {time_a}, B 타임: {time_b}")

                    # DataFrame 내용을 테이블 형식으로 출력
                    print_table(df)

                    # 기록 후 팀 번호 증가 및 타이머 결과 초기화
                    team_number += 1
                    time_a = None
                    time_b = None

except KeyboardInterrupt:
    print("프로그램 종료")

finally:
    ser.close()
